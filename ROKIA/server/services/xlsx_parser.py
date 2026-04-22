import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

# Column mapping for the ticket file (0-indexed)
# A=0: DIT no interne (ticket number)
# C=2: DIT Etat
# N=13: Cause
# P=15: Délai contractuel
# R=17: Contrat
# T=19: Description
# U=20: Résolution
COL_TICKET_NUMBER = 0    # A
COL_DIT_ETAT = 2         # C
COL_CAUSE = 13           # N
COL_DELAY = 15           # P
COL_CONTRACT = 17        # R
COL_DESCRIPTION = 19     # T
COL_RESOLUTION = 20      # U


def _safe_str(value) -> str:
    """Safely convert a cell value to string."""
    if value is None:
        return ""
    return str(value).strip()


def parse_reference(file_path: str) -> Dict[str, Any]:
    """Parse the ROKIA reference file.
    
    Returns:
        {
            'categories': ['cause1', 'cause2', ...],
            'contracts': [
                {'name': '...', 'options': '...', 'delay': '...', 'covered_elements': '...'},
                ...
            ]
        }
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    result = {'categories': [], 'contracts': []}

    # Parse CATEGORIES sheet
    if 'CATEGORIES' in wb.sheetnames:
        ws = wb['CATEGORIES']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                cat = _safe_str(row[0])
                if cat:
                    result['categories'].append(cat)
    else:
        logger.warning("Sheet 'CATEGORIES' not found in reference file")

    # Parse CONTRATS sheet
    if 'CONTRATS' in wb.sheetnames:
        ws = wb['CONTRATS']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                contract = {
                    'name': _safe_str(row[0]),
                    'options': _safe_str(row[1]) if len(row) > 1 else '',
                    'delay': _safe_str(row[2]) if len(row) > 2 else '',
                    'covered_elements': _safe_str(row[3]) if len(row) > 3 else '',
                }
                if contract['name']:
                    result['contracts'].append(contract)
    else:
        logger.warning("Sheet 'CONTRATS' not found in reference file")

    wb.close()

    logger.info(
        f"Reference parsed: {len(result['categories'])} categories, "
        f"{len(result['contracts'])} contracts"
    )
    return result


def parse_tickets(file_path: str) -> List[Dict[str, Any]]:
    """Parse the ticket file and return a list of ticket rows.
    
    Returns a list of dicts, one per row in the file.
    Multiple rows can share the same ticket_number.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    tickets = []
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        row_count += 1
        # Pad row to at least 21 columns
        padded = list(row) + [None] * max(0, 21 - len(row))

        ticket = {
            'ticket_number': _safe_str(padded[COL_TICKET_NUMBER]),
            'dit_etat': _safe_str(padded[COL_DIT_ETAT]),
            'old_category': _safe_str(padded[COL_CAUSE]),
            'old_delay': _safe_str(padded[COL_DELAY]),
            'old_contract': _safe_str(padded[COL_CONTRACT]),
            'description': _safe_str(padded[COL_DESCRIPTION]),
            'resolution': _safe_str(padded[COL_RESOLUTION]),
        }
        tickets.append(ticket)

    wb.close()
    logger.info(f"Parsed {row_count} ticket rows")
    return tickets


def group_tickets_by_number(tickets: List[Dict]) -> Dict[str, Dict]:
    """Group ticket rows by ticket number.
    
    For each unique ticket number, merge descriptions and resolutions
    from all rows to provide full context to the AI.
    """
    grouped = {}

    for t in tickets:
        num = t['ticket_number']
        if not num:
            continue

        if num not in grouped:
            grouped[num] = {
                'ticket_number': num,
                'dit_etat': t['dit_etat'],
                'old_category': t['old_category'],
                'old_delay': t['old_delay'],
                'old_contract': t['old_contract'],
                'descriptions': [],
                'resolutions': [],
            }

        if t['description']:
            grouped[num]['descriptions'].append(t['description'])
        if t['resolution']:
            grouped[num]['resolutions'].append(t['resolution'])

    # Merge descriptions and resolutions into single strings
    for num, data in grouped.items():
        data['merged_description'] = "\n---\n".join(
            list(dict.fromkeys(data['descriptions']))  # deduplicate
        )
        data['merged_resolution'] = "\n---\n".join(
            list(dict.fromkeys(data['resolutions']))
        )

    return grouped


def generate_result_xlsx(
    results: List[Dict],
    recurring_issues: List[Dict],
    client_name: str = "",
    copil_date: str = ""
) -> bytes:
    """Generate the result Excel file with two sheets.
    
    Sheet 1: Ticket results with recategorization
    Sheet 2: Recurring issues
    """
    wb = openpyxl.Workbook()

    # --- Styles ---
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    yes_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    yes_font = Font(name='Calibri', color='166534', bold=True)
    no_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    no_font = Font(name='Calibri', color='6B7280')

    # ========== Sheet 1: Results ==========
    ws1 = wb.active
    ws1.title = "Résultats"

    # Title row
    ws1.merge_cells('A1:I1')
    title_cell = ws1['A1']
    title_cell.value = f"ROKIA — Recatégorisation pour {client_name} — COPIL {copil_date}"
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='4F46E5')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = [
        'N° Ticket', 'DIT No Interne', 'DIT État',
        'Ancienne Catégorie', 'Nouvelle Catégorie',
        'Contrat', 'Délai Contractuel',
        'Recatégorisé', 'Justification IA'
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, result in enumerate(results, 4):
        was_recat = result.get('was_recategorized', False)
        row_data = [
            result.get('ticket_number', ''),
            result.get('dit_no_interne', ''),
            result.get('dit_etat', ''),
            result.get('old_category', ''),
            result.get('new_category', ''),
            result.get('new_contract', ''),
            result.get('new_delay', ''),
            'Oui' if was_recat else 'Non',
            result.get('ai_reasoning', ''),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

        # Style the "Recatégorisé" column
        recat_cell = ws1.cell(row=row_idx, column=8)
        if was_recat:
            recat_cell.fill = yes_fill
            recat_cell.font = yes_font
        else:
            recat_cell.fill = no_fill
            recat_cell.font = no_font
        recat_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    col_widths = [15, 18, 12, 25, 25, 20, 18, 15, 40]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    # Auto-filter
    if results:
        ws1.auto_filter.ref = f"A3:I{3 + len(results)}"

    # ========== Sheet 2: Recurring Issues ==========
    ws2 = wb.create_sheet("Problématiques récurrentes")

    headers2 = ['Catégorie / Cause', 'Nombre d\'occurrences', 'Tickets concernés', 'Détail']
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, issue in enumerate(recurring_issues, 2):
        row_data = [
            issue.get('category', ''),
            issue.get('count', 0),
            issue.get('tickets', ''),
            issue.get('detail', ''),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    col_widths2 = [30, 20, 30, 50]
    for i, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
